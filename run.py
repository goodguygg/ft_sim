import pandas as pd
from parts.utilities.utils import * 
from cadCAD.engine import ExecutionMode, ExecutionContext,Executor
from config import exp
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import json
from openpyxl.chart import LineChart, Reference, BarChart, AreaChart, series
from openpyxl.styles import PatternFill


def run():
    '''
    Definition:
    Run simulation
    '''
    # Single
    exec_mode = ExecutionMode()
    local_mode_ctx = ExecutionContext(context=exec_mode.local_mode)

    simulation = Executor(exec_context=local_mode_ctx, configs=exp.configs)
    raw_system_events, tensor_field, sessions = simulation.execute()
    # Result System Events DataFrame
    df = pd.DataFrame(raw_system_events)
 
    return df

def postprocessing(df):
    json_data = df.to_json(orient='records', indent=4)

    with open('data.json', 'w') as file:
        file.write(json_data)
    
    # Pool balances and ratios
    # Amount of traders and providers
    # Cumulative PnL for traders, max and min PnL for traders (per market)
    # Cumulative apy for liquidity providers (per market)
    # Open interest (long and short supplies per market)
    # Volumes
    # Number of liquidations
    # Fees collected
    # Treasury balance

    # track pnl of the pool as metric
    # Initialize an empty list to store each timestep data
    data = []
    # Loop through each row in the dataframe
    for _, row in df.iterrows():
        traders = row['traders']
        liquidity_providers = row['liquidity_providers']
        pools = row['pools']
        # treasury = row['treasury']
        liquidations = row['liquidations']
        num_of_longs = row['num_of_longs']
        num_of_shorts = row['num_of_shorts']
        num_of_swaps = row['num_of_swaps']

        nominal_exposure_btc = 0
        nominal_exposure_eth = 0
        nominal_exposure_sol = 0
        for trader in traders.values():
            # print(trader)
            if 'BTC' in trader['positions_long'] and 'BTC' in trader['positions_short']:
                # print(f"nom exp btc {trader['positions_long']['BTC']['quantity']} + {trader['positions_short']['BTC']['quantity']}")
                nominal_exposure_btc += trader['positions_long']['BTC']['quantity'] - trader['positions_short']['BTC']['quantity']
            elif 'BTC' in trader['positions_long']: 
                nominal_exposure_btc += trader['positions_long']['BTC']['quantity']
            elif 'BTC' in trader['positions_short']:
                nominal_exposure_btc -= trader['positions_short']['BTC']['quantity']
            if 'ETH' in trader['positions_long'] and 'ETH' in trader['positions_short']:
                # print(f"nom exp eth {trader['positions_long']['ETH']['quantity']} + {trader['positions_short']['ETH']['quantity']}")
                nominal_exposure_eth += trader['positions_long']['ETH']['quantity'] - trader['positions_short']['ETH']['quantity']
            elif 'ETH' in trader['positions_long']:
                nominal_exposure_eth += trader['positions_long']['ETH']['quantity']
            elif 'ETH' in trader['positions_short']:
                nominal_exposure_eth -= trader['positions_short']['ETH']['quantity']
            if 'SOL' in trader['positions_long'] and 'SOL' in trader['positions_short']:
                # print(f"nom exp sol {trader['positions_long']['SOL']['quantity']} + {trader['positions_short']['SOL']['quantity']}")
                nominal_exposure_sol += trader['positions_long']['SOL']['quantity'] - trader['positions_short']['SOL']['quantity']
            elif 'SOL' in trader['positions_long']:
                nominal_exposure_sol += trader['positions_long']['SOL']['quantity']
            elif 'SOL' in trader['positions_short']:
                nominal_exposure_sol -= trader['positions_short']['SOL']['quantity']

        # Generate data for each row
        timestep_data = {
            'number_of_traders': len(traders),
            'number_of_liquidity_providers': len(liquidity_providers),
            'pool_lp_tokens': pools[0]['lp_shares'],
            'pool_balance_btc': pools[0]['holdings']['BTC'],
            'pool_balance_eth': pools[0]['holdings']['ETH'],
            'pool_balance_sol': pools[0]['holdings']['SOL'],
            'pool_balance_usdc': pools[0]['holdings']['USDC'],
            'pool_balance_usdt': pools[0]['holdings']['USDT'],
            'cum_pnl_traders': sum(trader['PnL'] for trader in traders.values()),
            'max_pnl_traders': max(trader['PnL'] for trader in traders.values()),
            'min_pnl_traders': min(trader['PnL'] for trader in traders.values()),
            # 'cum_apy_providers': sum(lp['yield'] for lp in liquidity_providers.values()),  # Assuming each LP has a 'yield' key
            'oi_long_btc': pools[0]['oi_long']['BTC'],
            'oi_long_eth': pools[0]['oi_long']['ETH'],
            'oi_long_sol': pools[0]['oi_long']['SOL'],
            'oi_short_btc': pools[0]['oi_short']['BTC'],
            'oi_short_eth': pools[0]['oi_short']['ETH'],
            'oi_short_sol': pools[0]['oi_short']['SOL'],
            'volume_btc': pools[0]['volume']['BTC'],
            'volume_eth': pools[0]['volume']['ETH'],
            'volume_sol': pools[0]['volume']['SOL'],
            'num_of_longs': num_of_longs,
            'num_of_shorts': num_of_shorts,
            'num_of_swaps': num_of_swaps,
            'number_of_liquidations': liquidations,
            'fees_collected_btc': pools[0]['total_fees_collected']['BTC'],
            'fees_collected_eth': pools[0]['total_fees_collected']['ETH'],
            'fees_collected_sol': pools[0]['total_fees_collected']['SOL'],
            'fees_collected_usdc': pools[0]['total_fees_collected']['USDC'],
            'fees_collected_usdt': pools[0]['total_fees_collected']['USDT'],
            'treasury_balance_btc': liquidity_providers['genesis']['liquidity']['BTC'],
            'treasury_balance_eth': liquidity_providers['genesis']['liquidity']['ETH'],
            'treasury_balance_sol': liquidity_providers['genesis']['liquidity']['SOL'],
            'treasury_balance_usdc': liquidity_providers['genesis']['liquidity']['USDC'],
            'treasury_balance_usdt': liquidity_providers['genesis']['liquidity']['USDT'],
            'pool_pnl_btc': pools[0]['open_pnl_long']['BTC'] + pools[0]['open_pnl_short']['BTC'], 
            'pool_pnl_eth': pools[0]['open_pnl_long']['ETH'] + pools[0]['open_pnl_short']['ETH'],
            'pool_pnl_sol': pools[0]['open_pnl_long']['SOL'] + pools[0]['open_pnl_short']['SOL'],
            'nominal_exposure_btc': nominal_exposure_btc,
            'nominal_exposure_eth': nominal_exposure_eth,
            'nominal_exposure_sol': nominal_exposure_sol,
        }
        
        # Append the timestep data to the list
        data.append(timestep_data)

    # Convert the list of timestep data into a DataFrame
    data_df = pd.DataFrame(data)

    return data_df

def to_xslx(df, name):

    wb = openpyxl.Workbook()
    sheet = wb.active
    rows = dataframe_to_rows(df)
    for r_idx, row in enumerate(rows, 1):
        for c_idx, value in enumerate(row, 1):
            try:
                value = float(value)
            except:
                pass
            sheet.cell(row=r_idx, column=c_idx, value=value)
    timestamps = df.shape[0]

    # Create a traction sheet
    trac_sheet = wb.create_sheet(title="Traction charts")
    # red code
    cells = [trac_sheet.cell(row=3, column=x) for x in range(21,24)]
    cells[0].value = "Key tested values"
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    [setattr(cell, 'fill', red_fill) for cell in cells]
    # green code
    cells = [trac_sheet.cell(row=5, column=x) for x in range(21,24)]
    cells[0].value = "Input controlled values"
    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    [setattr(cell, 'fill', green_fill) for cell in cells]
    # blue code
    cells = [trac_sheet.cell(row=7, column=x) for x in range(21,24)]
    cells[0].value = "Context values"
    blue_fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")
    [setattr(cell, 'fill', blue_fill) for cell in cells]

    trac_sheet['A1'] = "Amount of token lps"
    values = Reference(sheet, min_col=3, min_row=3, max_col=3, max_row=timestamps)
    chart = AreaChart()
    chart.add_data(values)
    chart.title = "number_of_liquidity_providers"
    chart.x_axis.title = "Hr"
    chart.y_axis.title = "LPs"
    # Change bar filling and line color 
    s = chart.series[0]
    s.graphicalProperties.line.solidFill = "00FF00"
    s.graphicalProperties.solidFill = "00FF00"
    trac_sheet.add_chart(chart, "A3")

    trac_sheet['L1'] = "Amount of traders"
    values = Reference(sheet, min_col=2, min_row=3, max_col=2, max_row=timestamps)
    chart = LineChart()
    chart.add_data(values)
    chart.title = "number_of_traders"
    chart.x_axis.title = "Hr"
    chart.y_axis.title = "Trads"
    # Change bar filling and line color 
    s = chart.series[0]
    s.graphicalProperties.line.solidFill = "00FF00"
    s.graphicalProperties.solidFill = "00FF00"
    trac_sheet.add_chart(chart, "L3")

    # Create a eth maxi sheet
    pool_sheet = wb.create_sheet(title="pool charts")

    pool_sheet['A1'] = "BTC pool size"
    values = Reference(sheet, min_col=5, min_row=3, max_col=5, max_row=timestamps)
    chart = AreaChart()
    chart.add_data(values)
    chart.title = "pool_balance_btc"
    chart.x_axis.title = "Hr"
    chart.y_axis.title = "BTC"
    # Change bar filling and line color 
    s = chart.series[0]
    s.graphicalProperties.line.solidFill = "FF0000"
    s.graphicalProperties.solidFill = "FF0000"
    pool_sheet.add_chart(chart, "A3")

    pool_sheet['J1'] = "ETH pool size"
    values = Reference(sheet, min_col=6, min_row=3, max_col=6, max_row=timestamps)
    chart = LineChart()
    chart.add_data(values)
    chart.title = "pool_balance_eth"
    chart.x_axis.title = "Hr"
    chart.y_axis.title = "ETH"
    pool_sheet.add_chart(chart, "J3")

    pool_sheet['S1'] = "SOL pool size"
    values = Reference(sheet, min_col=7, min_row=3, max_col=7, max_row=timestamps)
    chart = BarChart()
    chart.add_data(values)
    chart.title = "pool_balance_sol"
    chart.x_axis.title = "Hr"
    chart.y_axis.title = "SOL"
    # Change bar filling and line color 
    s = chart.series[0]
    s.graphicalProperties.line.solidFill = "00FF00"
    s.graphicalProperties.solidFill = "00FF00"
    pool_sheet.add_chart(chart, "S3")

    pool_sheet['A18'] = "USDC pool size"
    values = Reference(sheet, min_col=8, min_row=3, max_col=8, max_row=timestamps)
    chart = BarChart()
    chart.add_data(values)
    chart.title = "pool_balance_usdc"
    chart.x_axis.title = "Hr"
    chart.y_axis.title = "USDC"
    # Change bar filling and line color 
    s = chart.series[0]
    s.graphicalProperties.line.solidFill = "FF0000"
    s.graphicalProperties.solidFill = "FF0000"
    pool_sheet.add_chart(chart, "A20")

    pool_sheet['J18'] = "USDT pool size"
    values = Reference(sheet, min_col=9, min_row=3, max_col=9, max_row=timestamps)
    chart = AreaChart()
    chart.add_data(values)
    chart.title = "pool_balance_usdt"
    chart.x_axis.title = "Hr"
    chart.y_axis.title = "USDT"
    pool_sheet.add_chart(chart, "J20")

    # eth_sheet['S18'] = "Trade volume of the eth maxi tokens"
    # timestamps = df.shape[0]
    # values = Reference(sheet, min_col=11, min_row=3, max_col=11, max_row=timestamps)
    # chart = LineChart()
    # chart.add_data(values)
    # chart.title = "eth_volume"
    # chart.x_axis.title = "Day"
    # chart.y_axis.title = "ETHM"
    # # Change bar filling and line color 
    # s = chart.series[0]
    # s.graphicalProperties.line.solidFill = "00FF00"
    # s.graphicalProperties.solidFill = "00FF00"
    # eth_sheet.add_chart(chart, "S20")

    # eth_sheet['A35'] = "Trade volume of the eth maxi tokens in USD"
    # timestamps = df.shape[0]
    # values = Reference(sheet, min_col=12, min_row=3, max_col=12, max_row=timestamps)
    # chart = LineChart()
    # chart.add_data(values)
    # chart.title = "eth_pvolume"
    # chart.x_axis.title = "Day"
    # chart.y_axis.title = "$"
    # eth_sheet.add_chart(chart, "A37")

    # eth_sheet['J35'] = "Volume of arbitrage trades"
    # timestamps = df.shape[0]
    # values = Reference(sheet, min_col=13, min_row=3, max_col=13, max_row=timestamps)
    # chart = LineChart()
    # chart.add_data(values)
    # chart.title = "eth_arbitrage_volume"
    # chart.x_axis.title = "Day"
    # chart.y_axis.title = "$"
    # # Change bar filling and line color 
    # s = chart.series[0]
    # s.graphicalProperties.line.solidFill = "00FF00"
    # s.graphicalProperties.solidFill = "00FF00"
    # eth_sheet.add_chart(chart, "J37")

    # eth_sheet['S35'] = "Spread between trades up and down"
    # timestamps = df.shape[0]
    # values = Reference(sheet, min_col=14, min_row=3, max_col=14, max_row=timestamps)
    # chart = LineChart()
    # chart.add_data(values)
    # chart.title = "eth_arbitrage_spread"
    # chart.x_axis.title = "Day"
    # chart.y_axis.title = "$"
    # eth_sheet.add_chart(chart, "S37")

    # eth_sheet['A52'] = "Balance of the eth maxi SAV"
    # timestamps = df.shape[0]
    # values = Reference(sheet, min_col=15, min_row=3, max_col=15, max_row=timestamps)
    # chart = LineChart()
    # chart.add_data(values)
    # chart.title = "sav_balance_eth"
    # chart.x_axis.title = "Day"
    # chart.y_axis.title = "$"
    # # Change bar filling and line color 
    # s = chart.series[0]
    # s.graphicalProperties.line.solidFill = "FF0000"
    # s.graphicalProperties.solidFill = "FF0000"
    # eth_sheet.add_chart(chart, "A54")


    wb.save(f'{name}.xlsx')

def main():
    '''
    Definition:
    Run simulation and extract metrics
    '''
    df = run()
    df = postprocessing(df)
    to_xslx(df, 'run') 
    df = df[df.index % 2 != 0]
    df = df.reset_index(drop=True)
    to_xslx(df, 'run_merged') 
    return df

if __name__ == '__main__':
    main()